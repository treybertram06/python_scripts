import openpyxl
from datetime import datetime

name = input('Enter your name: ')

# Load workbook
wb = openpyxl.load_workbook('JADE-Report.xlsx')

# Select sheet
sheet = wb['Sheet1']

# Get today's date
today = datetime.today().strftime('%A, %B %d, %Y')

# Find the next empty row in column B (Date of Class)
next_row = 1
while sheet.cell(row=next_row, column=2).value is not None:
    next_row += 1

# Fill in the report
sheet.cell(row=next_row, column=1, value=75)  # This class
sheet.cell(row=next_row, column=4, value=75)  # Engaged!
sheet.cell(row=next_row, column=11, value='Used python script to automatically input 75 minutes of engaged time')  # Other
sheet.cell(row=next_row, column=12, value=(name + " worked very hard.") ) 

# Save workbook
wb.save('JADE-Report.xlsx')

print('Report has been updated for', today + '.')